from __future__ import annotations

import glob
import os
import shutil
import subprocess
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from scipy.io import loadmat, savemat
from scipy.sparse.linalg import lsqr as scipy_lsqr

from paths import SRC_DIR, poseidon_exe_path

np.set_printoptions(linewidth=160, threshold=10000)
MODULE_DIR = SRC_DIR
PROJECT_ROOT = MODULE_DIR.parent


def resolve_path(path_like: str | os.PathLike) -> Path:
    path = Path(path_like)
    if path.exists():
        return path
    module_relative = MODULE_DIR / path
    if module_relative.exists():
        return module_relative
    return path


def load_numeric_excel(path_like: str | os.PathLike) -> np.ndarray:
    path = resolve_path(path_like)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    nrows = len(rows)
    ncols = max((len(r) for r in rows), default=0)
    arr = np.full((nrows, ncols), np.nan, dtype=float)
    for i, row in enumerate(rows):
        for j, value in enumerate(row):
            if value is None:
                continue
            try:
                arr[i, j] = float(value)
            except Exception:
                arr[i, j] = np.nan
    wb.close()
    return arr


def _normalize_2d(data) -> np.ndarray:
    if np.isscalar(data):
        return np.array([[data]])
    arr = np.asarray(data)
    if arr.ndim == 0:
        return arr.reshape(1, 1)
    if arr.ndim == 1:
        return arr.reshape(-1, 1)
    return arr


def xlswrite_compat(path_like: str | os.PathLike, data, sheet_name: str = 'Sheet1') -> None:
    path = Path(path_like)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        try:
            wb = load_workbook(path)
        except Exception:
            wb = Workbook()
            wb.active.title = 'Sheet1'
    else:
        wb = Workbook()
        wb.active.title = 'Sheet1'
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=str(sheet_name).strip())
    arr = _normalize_2d(data)
    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            value = arr[i, j]
            if isinstance(value, np.generic):
                value = value.item()
            ws.cell(row=i + 1, column=j + 1, value=value)
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1 and wb['Sheet'].max_row == 1 and wb['Sheet'].max_column == 1 and wb['Sheet']['A1'].value is None:
        wb.remove(wb['Sheet'])
    wb.save(path)
    wb.close()


def save_mat_compat(path_like: str | os.PathLike, **kwargs) -> None:
    savemat(path_like, kwargs)


def load_mat_compat(path_like: str | os.PathLike):
    return loadmat(path_like, squeeze_me=True, struct_as_record=False)


def matlab_print(name: str, value) -> None:
    print(f"\n{name} =\n")
    print(value)
    print()


def find_poseidon_folders(date_path: str | os.PathLike) -> list[str]:
    date_dir = Path(date_path)
    return sorted([p.name for p in date_dir.glob('*hydrol*dep*') if p.is_dir()])


def mkdir_warning(path_like: str | os.PathLike) -> None:
    path = Path(path_like)
    if path.exists():
        print('Warning: Directory already exists.')
    else:
        path.mkdir(parents=True, exist_ok=True)


def copy_matching(pattern: str, destination: str | os.PathLike) -> None:
    dest = Path(destination)
    dest.mkdir(parents=True, exist_ok=True)
    for match in sorted(glob.glob(pattern)):
        src = Path(match)
        target = dest / src.name
        if src.is_dir():
            shutil.copytree(src, target, dirs_exist_ok=True)
        else:
            shutil.copy2(src, target)


def copy_file(src: str | os.PathLike, dst: str | os.PathLike) -> None:
    src_p = Path(src)
    dst_p = Path(dst)
    dst_p.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_p, dst_p)


def matlab_lsqr(A, b, atol: float = 1e-8, iter_lim: int = 10000) -> np.ndarray:
    result = scipy_lsqr(A, b, atol=atol, btol=atol, iter_lim=iter_lim, show=False)
    x = result[0]
    itn = int(result[2])
    r1norm = float(result[3])
    denom = float(np.linalg.norm(np.asarray(b).ravel()))
    relres = 0.0 if denom == 0.0 else r1norm / denom
    print(f"lsqr converged at iteration {itn} to a solution with relative residual {relres:.3g}.")
    return x


def is_windows() -> bool:
    return os.name == 'nt'


def _poseidon_exe_path(exe_name: str = 'Poseidon4.exe') -> Path:
    return poseidon_exe_path()


def resolve_executable_command(exe_name: str = 'Poseidon4.exe') -> list[str]:
    exe_path = _poseidon_exe_path(exe_name)
    if is_windows():
        return [str(exe_path)]
    wine = shutil.which('wine')
    if wine:
        return [wine, str(exe_path)]
    raise RuntimeError(f'Cannot run {exe_path}')


def run_executable_in_dir(directory: str | os.PathLike, exe_name: str = 'Poseidon4.exe') -> None:
    folder_dir = Path(directory)
    exe_path = _poseidon_exe_path(exe_name)
    if not exe_path.exists():
        raise FileNotFoundError(f'Poseidon executable not found: {exe_path}')

    print()
    print(f"{folder_dir}>{exe_name}")
    result = subprocess.run(resolve_executable_command(exe_name), cwd=folder_dir)
    if result.returncode != 0:
        raise RuntimeError(f'{exe_name} failed in {folder_dir} with exit code {result.returncode}')
    print()
